import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import json
import re
import sys
import os

column_mapping = {
    "Id Number": "A",
    "RFID": "B",
    "Item Category": "C",
    "Item Description": "D",
    "Model": "E",
    "SWL Value": "F",
    "SWL Unit": "G",
    "SWL Note": "H",
    "Manufacturer": "I",
    "Certificate No": "J",
    "Location": "K",
    "Detailed Location ": "L",
    "Previous Inspection": "M",
    "Next Inspection Due Date": "N",
    "Fit For Purpose Y/N": "O",
    "Status": "P",
    "Provider Identification": "Q",
    "Errors": "R"
}

def preProcess(sheet_data, column_mapping):
    for header, column in column_mapping.items():
        cell = sheet_data[column + '2']
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Center align the text
        cell.border = Border(bottom=Side(border_style='thin'))  # Add a thin border at the bottom

        # Adjust column width to fit the header text
        column_width = max(len(header), 10)  # Set a minimum width of 10 characters
        sheet_data.column_dimensions[column].width = column_width

def insertData(extracted_data, sheet_data, error_pages):
    if extracted_data is None:
        print("Error: extracted_data is None")
        return

    row_idx = 3
    max_length = max(len(values) for values in extracted_data.values() if isinstance(values, list))

    # Process all columns except SWL separately
    for i in range(max_length):
        for name, values in extracted_data.items():
            if name == "SWL":
                continue  # Skip SWL, it will be processed separately

            column_name = column_mapping.get(name)
            if column_name and isinstance(values, list) and i < len(values):
                val = values[i]
                # print(f"Processing column '{name}' with value '{val}' of type '{type(val)}'")
                if not isinstance(val, str):  # Convert any non-string value to string
                    val = str(val)
                if column_name == "M" or column_name == "N":
                    val = dateProcess(val)
                if column_name == "E":
                    modelTuple = modelProcess(val)
                    if modelTuple:
                        val = modelTuple[0]
                        manu = modelTuple[1]
                        sheet_data.cell(row=row_idx + i, column=ord("I") - 64, value=manu)
                if 'Bounding box is outside the page bounds.' in val:
                    error_pages.append(f"Page {row_idx + i - 3}, Column {name}")
                sheet_data.cell(row=row_idx + i, column=ord(column_name) - 64, value=val)

    # Process SWL separately
    if "SWL" in extracted_data:
        swlValue = extracted_data["SWL"]
        swlProcess(sheet_data, swlValue, row_idx, error_pages)

def insertError(workbook, error_pages):
    if error_pages is None:
        print("Error: error_pages is None")
        return

    sheet_errors = workbook.create_sheet(title="Errors")  # Create a new worksheet
    sheet_errors.append(["No", "Error"])  # Write column headers

    idx = 2
    for error in error_pages:
        print(f"Processing error value '{error}'")
        sheet_errors.cell(row=idx, column=ord("A") - 64, value=idx - 1)
        sheet_errors.cell(row=idx, column=ord("B") - 64, value=error)
        idx += 1

def create_excel(extracted_data: dict, filename: str, client: str, page_errors: dict):
    try:
        print("<--------------Creating new excel------------------>")
        workbook = openpyxl.Workbook()  # Create a new Workbook

        # Create a sheet for extracted data
        sheet_data = workbook.active
        sheet_data.title = "Extraction Data"  # Set sheet name

        sheet_data['A1'] = "Rig-Ware import v2"
        sheet_data['B1'] = client
        sheet_data['C1'] = "CreateLocations=No"

        # Write column headers for extracted data sheet
        preProcess(sheet_data, column_mapping)
        
        # List to hold error pages
        error_pages = []

        # Insert data from json 
        insertData(extracted_data, sheet_data, error_pages)

        # Insert error information
        insertError(workbook, error_pages)

        workbook.save(filename)  # Save the workbook with the provided filename
        workbook.close()
        print("<-------------- Excel created successfully ------------------>")
    except Exception as e:
        print(f"An error occurred in excel creation: {e}")

def modelProcess(modelData):
    print(f"Processing model data '{modelData}' of type '{type(modelData)}'")
    if not isinstance(modelData, str):
        modelData = str(modelData)  # Ensure modelData is a string
    workbook = load_workbook("database/Full_list_of_Manufacturers_and_Models.xlsx")
    model_sheet = workbook['Model']
    for row in model_sheet.iter_rows(min_row=2, values_only=True):
        keyword = row[0]
        value = row[1]
        if isinstance(keyword, int):
            keyword = str(keyword)
        if keyword in modelData:
            return keyword, value
    return None

def manuProcess(manuData):
    print(f"Processing manufacturer data '{manuData}' of type '{type(manuData)}'")
    if not isinstance(manuData, str):
        manuData = str(manuData)  # Ensure manuData is a string
    workbook = load_workbook("database/Full_list_of_Manufacturers_and_Models.xlsx")
    manufacturer_sheet = workbook['Manufacture']
    for row in manufacturer_sheet.iter_rows(min_row=2, values_only=True):
        keyword = row[0]
        value = row[1]
        if isinstance(keyword, int):
            keyword = str(keyword)
        if keyword in manuData:
            return value
    return None

def dateProcess(val):
    print(f"Processing date value '{val}' of type '{type(val)}'")
    if not isinstance(val, str):
        val = str(val)
    pattern = r'\b\d{2}[-/](?:\d{2}|[A-Za-z]{3})[-/]\d{4}\b|[A-Za-z]+(?:\s+[A-Za-z]+)*'
    val = re.findall(pattern, val)
    if len(val) != 0:
        val = val[0]
    return val


def swlProcess(sheet_data, swlValue, start_row_idx, error_pages):
    if swlValue is None:
        print("Error: swlValue is None")
        return
    
    valPattern = r'\d+(?:\.\d+)?'
    unitPattern = r'[a-zA-Z]+'
    notePattern = r'[\s\n]+(\S.*)'
    
    for i, val in enumerate(swlValue):
        try:
            current_row_idx = start_row_idx + i
            print(f"Processing SWL value '{val}' of type '{type(val)}' at row {current_row_idx}")
            if val is None:
                raise ValueError("SWL value is None")
            if not isinstance(val, str):  # Convert any non-string value to string
                val = str(val)
            swlVal_matches = re.findall(valPattern, val)
            if not swlVal_matches:
                raise ValueError(f"No SWL value match found in '{val}'")
            swlVal = swlVal_matches[0]
            sheet_data.cell(row=current_row_idx, column=ord("F") - 64, value=swlVal)
            
            swlUnit_matches = re.findall(unitPattern, val)
            if not swlUnit_matches:
                raise ValueError(f"No SWL unit match found in '{val}'")
            swlUnit = swlUnit_matches[0]
            sheet_data.cell(row=current_row_idx, column=ord("G") - 64, value=swlUnit)
            
            swlNote_matches = re.findall(notePattern, val)
            if len(swlNote_matches) > 0 and swlNote_matches[0] != swlUnit:
                swlNote = swlNote_matches[0]
                sheet_data.cell(row=current_row_idx, column=ord("H") - 64, value=swlNote)
        except Exception as e:
            error_message = f"Error processing SWL at Page {current_row_idx}: {e}"
            print(error_message)
            error_pages.append(error_message)


def jsonProcess(filepath):
    if os.path.isfile(filepath):
        try:
            with open(filepath, 'r') as file:
                data = json.load(file)
            return data
        except Exception as e:
            print(f"Error loading JSON data from {filepath}: {e}")
            return None
    else:
        return filepath  # Return the error message string directly

if __name__ == "__main__":
    if len(sys.argv) != 5:
        print("Usage: python excel_management.py <input JSON file> <output Excel file> <Client Name> <Error JSON file>")
        sys.exit(1)

    data = jsonProcess(sys.argv[1])
    page_errors = jsonProcess(sys.argv[4])
    create_excel(data, sys.argv[2], sys.argv[3], page_errors)